import openpyxl
from Data_Driven_Framework import XL_Utils
from selenium import webdriver

file=r"C:\Users\ASEEAZ\Desktop\selenuim-python\newxlutil.xlsx"
bk=openpyxl.load_workbook(file)
s=bk.active
driver=webdriver.Chrome()
rowcount=s.max_row

print(rowcount)
colcount=s.max_column
print(colcount)

for curr_row in range(1,rowcount+1):
    username=XL_Utils.readData(s,0,curr_row,1)
    password=XL_Utils.readData(s,0,curr_row,2)
    print(("user :" + username + " / password" + password))
    driver.get("https://mail.rediff.com/cgi-bin/login.cgi")

    driver.find_element_by_id("login1").send_keys(username)
    driver.find_element_by_id("password").send_keys(password)
    driver.find_element_by_name("proceed").click()
    if driver.title == "Rediffmail":
        XL_Utils.WriteData(s,"sheet1",curr_row,3,"testpassed")
    else:
        XL_Utils.WriteData(s, "sheet1", curr_row, 3, "testfailed")
bk.save(file)